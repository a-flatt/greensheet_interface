import psycopg2
import openpyxl
from openpyxl.utils import get_column_letter
import greensheet_utils as utils
import greensheet_queries
from openpyxl.formula.translate import Translator

def row_adjust(ws, item_list, row_start, row_finish):
	
	cur_rowrange = find_row_index(ws, row_start, row_finish)
	row_count = cur_rowrange[1] - cur_rowrange[0] - 1
	req_rows = len(item_list)
	rows_to_add = req_rows - row_count
	rows_to_delete = row_count - req_rows

	if req_rows > row_count:
		ws.insert_rows(cur_rowrange[1] - 2, rows_to_add)
		copy_row_to_rows(ws, cur_rowrange[0] + 3, cur_rowrange[0] + 3 + rows_to_add, 12)
	elif req_rows < row_count:
		ws.delete_rows(cur_rowrange[0] + 1, rows_to_delete)
	else:
		return

def find_row_index(ws, start_ref, finish_ref):

	"""
	Use cell values (I.E 'CS', 'CF') to return row coordinate. 
	'CS' will always be at A12, however, 'CF', 'LS' and 'LF' will change as rows are added. 
	"""

	for row in ws['A']:
		if row.value == start_ref:
			row_start = row.row
		elif row.value == finish_ref:
			row_finish = row.row
	return [row_start, row_finish]

def copy_row_to_rows(ws, start_row, finish_row, src_row):

	for row in range(start_row, finish_row + 1):
		copy_row(ws, row, src_row)

def copy_row(ws, tgt_row, src_row):

	tgt_range = '{}:{}'.format(tgt_row, tgt_row)

	for col in range(1, utils.num_styled_cols(ws, tgt_range)):
		ws[utils.formatter(col, tgt_row)]._style = ws[utils.formatter(col, src_row)]._style

def insert(ws, item_list, start_ref, finish_ref):

	r = find_row_index(ws, start_ref, finish_ref)[0] + 1
	for row in item_list:
		c = 1
		for column in row:
			ws.cell(row = r, column = c).value = column
			c +=1
		r +=1
	
def main():

	# Open and activate workbook to write to. 

	wb = openpyxl.load_workbook('testproject.xlsx')
	ws = wb.active
	
	# Find values of paramaters for use in PostGRES query. 
	income_id = ws['A2'].value
	cost_id = ws['A1'].value
	
	# Retrieve lists from Postgres database. 
	cost_list = greensheet_queries.fetchR2([1, 59999], cost_id)
	labour_list = greensheet_queries.fetchR2([60000, 69999], cost_id)
	income_list = greensheet_queries.fetchBOQ(income_id)

	# Add income to allocated cell. 
	ws['C5'].value = utils.contract_value(income_list)

	# Adjust number of rows in spreadws to match len() of lists. 
	row_adjust(ws, cost_list, 'CS', 'CF')
	row_adjust(ws, labour_list, 'LS', 'LF')
	# copy_formulas(ws, cost_list, 'CS', 'CF')
	# copy_formulas(ws, labour_list, 'LS', 'LF')
	
	# Insert retrieved values from database into spreadws. 
	insert(ws, cost_list, 'CS', 'CF')
	insert(ws, labour_list, 'LS', 'LF')

	# Reformat cells, colours etc. 
	utils.reformat_sheet(ws)

	wb.save('testproject1.xlsx') 

main()
