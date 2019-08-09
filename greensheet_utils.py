import openpyxl
from openpyxl.utils import get_column_letter

def reformat_sheet(worksheet):

	for row in range(1, worksheet.max_row + 1):
		worksheet.row_dimensions[row].height = 21.0

def formatter(col="", row=""):

	return get_column_letter(col) + str(row)

def num_styled_cols(worksheet, tgt_range):

	return len([cell._style for cell in worksheet[tgt_range]])

def add_totals(test_list):

	for row in test_list:
		row.append(row[2] * row[3])
	return test_list

def contract_value(boq_list):

	new_list = add_totals(boq_list)
	contract_value = 0
	for row in new_list:
		contract_value += row[4]
	return contract_value
